#!/usr/bin/env python3
"""
database_extract_source_data.py

Extracts, filters, and merges the primary-literature supplementary tables that
underlie CpG_Brain_Peripheral_Correlation_Database.xlsx into two master CSVs:

  - cpg_brain_peripheral_correlation_database.csv  (9,127 CpG x tissue-pair rows)
  - gene_level_summary.csv                          (815 genes, one row per gene)

This step reads the raw journal supplementary files and applies the
|r| / |rho| > 0.6 threshold. 
It is upstream of build_cpg_database.py, which then assembles 
the final formatted .xlsx from these two CSVs.

INPUT FILES (must be downloaded from the two source papers first):

  1. essex_supp.xlsx
     Hannon E, Lunnon K, Schalkwyk L, Mill J. "Interindividual methylomic
     variation across blood, cortex, and cerebellum: implications for
     epigenetic studies of neurological and neuropsychiatric phenotypes."
     Epigenetics 2015;10(11):1024-1032. doi:10.1080/15592294.2015.1100786
     Article: https://pmc.ncbi.nlm.nih.gov/articles/PMC4844197/
     Supplemental material (zip, ~5.1MB), listed on the article page as
     "kepi-10-11-1100786-s001.zip" — download and unzip, then rename/convert
     the combined supplementary-tables workbook to essex_supp.xlsx with sheets:
       - "Supplementary Table 2"  (450K manifest annotation: CHR, MAPINFO,
         UCSC_REFGENE_NAME, UCSC_REFGENE_GROUP, RELATION_TO_UCSC_CPG_ISLAND)
       - "Supplementary Table 6"  (blood-brain r for the "low-variance" probe
         set — checked here for completeness, contributes 0 qualifying rows)
       - "Supplementary Table 7"  (blood-brain r for the "highly-variable"
         probe set — source of essentially all qualifying blood-brain rows)

  2. sommerer_supp1.xlsx
     Sommerer Y, Ohlei O, Dobricic V, et al. "A correlation map of
     genome-wide DNA methylation patterns between paired human brain and
     buccal samples." Clinical Epigenetics 2022;14:118.
     doi:10.1186/s13148-022-01357-w
     Article: https://clinicalepigeneticsjournal.biomedcentral.com/articles/10.1186/s13148-022-01357-w
     Additional file 1 (supplementary tables Excel workbook), with sheets:
       - "S2"  (CpG_ID, CHR, MAPINFO, Rho_publication, Rho_reanalysis, ...)
       - "S6"  (CpG, Spearman R, p-value, q-value, buccal/blood/brain_mQTL,
         CoRSIV, Braun_et_al — the "top q<0.05 correlated probes" table)

Run: python3 extract_source_data.py
"""

import csv
import openpyxl

WORKSPACE = "/home/user/workspace"
ESSEX_XLSX = f"{WORKSPACE}/essex_supp.xlsx"
SOMMERER_XLSX = f"{WORKSPACE}/sommerer_supp1.xlsx"

ESSEX_HIGH_CORR_CSV = f"{WORKSPACE}/essex_high_corr.csv"
BUCCAL_HIGH_CORR_CSV = f"{WORKSPACE}/buccal_brain_high_corr.csv"
MASTER_CSV = f"{WORKSPACE}/cpg_brain_peripheral_correlation_database.csv"
GENE_SUMMARY_CSV = f"{WORKSPACE}/gene_level_summary.csv"

CORR_THRESHOLD = 0.6

HANNON_SOURCE = "Hannon et al. 2015, Epigenetics 10(11):1024-32 (Blood Brain DNA Methylation Comparison Tool)"
HANNON_URL = "https://epigenetics.essex.ac.uk/bloodbrain/ ; https://pmc.ncbi.nlm.nih.gov/articles/PMC4844197/"
HANNON_TOOL = "Blood Brain DNA Methylation Comparison Tool (Univ. Exeter/Essex)"

SOMMERER_SOURCE = "Sommerer et al. 2022, Clinical Epigenetics 14:118 (MADRC buccal-brain correlation map)"
SOMMERER_URL = "http://www.liga.uni-luebeck.de/buccal_brain_correlation_results/ ; https://clinicalepigeneticsjournal.biomedcentral.com/articles/10.1186/s13148-022-01357-w"
SOMMERER_TOOL = "MADRC Buccal-Brain Correlation Map"

# Illumina UCSC_RefGene_Group tokens that indicate a promoter-proximal region.
# Any probe whose gene annotation includes at least one of these (even if it
# also has Body/3'UTR tokens for other transcripts) is classified as
# "Promoter-associated" — promoter status takes priority over gene-body status.
PROMOTER_TOKENS = {"TSS200", "TSS1500", "5'UTR", "1stExon"}
BODY_TOKENS = {"Body", "3'UTR"}


def classify_context(raw_group):
    """Map a raw (possibly multi-transcript, semicolon-joined) UCSC_RefGene_Group
    string to one of: Promoter-associated / Gene body/intragenic / Unannotated/Intergenic."""
    if not raw_group:
        return "Unannotated/Intergenic"
    tokens = {t.strip() for t in raw_group.split(";") if t.strip()}
    if tokens & PROMOTER_TOKENS:
        return "Promoter-associated"
    if tokens & BODY_TOKENS:
        return "Gene body/intragenic"
    return "Unannotated/Intergenic"


# ===========================================================================
# STEP 1 — Build the shared Illumina 450K manifest annotation lookup from
# Hannon et al.'s Supplementary Table 2. This is reused for BOTH the
# blood-brain (Essex) rows and the buccal-brain (Sommerer) rows below, since
# the two studies' probe sets substantially overlap and only the Essex table
# carries gene / genomic-context / CpG-island annotation.
# ===========================================================================
print("Loading Illumina manifest annotation from essex_supp.xlsx Supplementary Table 2 ...")
wb_essex = openpyxl.load_workbook(ESSEX_XLSX, read_only=True)
annotation_lookup = {}  # probe -> (gene, group, relation_to_island)
ws_anno = wb_essex["Supplementary Table 2"]
for row in ws_anno.iter_rows(min_row=3, values_only=True):
    probe = row[0]
    if not probe:
        continue
    gene = row[6] or ""
    group = row[7] or ""
    relation = row[9] or ""
    annotation_lookup[probe] = (gene, group, relation)
print(f"  Annotated probes in lookup: {len(annotation_lookup):,}")


# ===========================================================================
# STEP 2 — Extract Hannon et al. 2015 blood-brain correlations
# (Supplementary Tables 6 + 7), filter to |r| > 0.6, join with the manifest
# lookup, and write the long-format intermediate CSV.
# ===========================================================================
REGION_COLS = {"PFC": 1, "EC": 4, "STG": 7, "CER": 10}  # 0-based col index of the 'r' value

essex_rows = []
for table_name in ["Supplementary Table 6", "Supplementary Table 7"]:
    ws = wb_essex[table_name]
    n_qualifying = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        probe = row[0]
        if not probe:
            continue
        gene, group, relation = annotation_lookup.get(probe, ("", "", ""))
        chr_val, pos_val = None, None
        # CHR/MAPINFO are not present in Tables 6/7 themselves — pull from the
        # same manifest lookup used for gene annotation (Table 2 also carries
        # CHR/MAPINFO in columns 4/5, so extend the lookup inline here).
        for region, col in REGION_COLS.items():
            r = row[col]
            if r is None or abs(r) <= CORR_THRESHOLD:
                continue
            n_qualifying += 1
            essex_rows.append({
                "probe": probe, "region": region, "r": r,
                "chr": chr_val, "mapinfo": pos_val,
                "gene": gene, "group": group, "relation": relation,
                "table": table_name,
            })
    print(f"  {table_name}: {n_qualifying:,} qualifying (probe, region) pairs with |r| > {CORR_THRESHOLD}")

# CHR/MAPINFO come from the Table 2 manifest rows too — build a second lookup
# and backfill, since Tables 6/7 only carry ProbeID + correlation stats.
chrpos_lookup = {}
for row in ws_anno.iter_rows(min_row=3, values_only=True):
    probe = row[0]
    if probe:
        chrpos_lookup[probe] = (row[4], row[5])  # CHR, MAPINFO
for r in essex_rows:
    chr_val, pos_val = chrpos_lookup.get(r["probe"], (None, None))
    r["chr"], r["mapinfo"] = chr_val, pos_val

with open(ESSEX_HIGH_CORR_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["probe", "region", "r", "chr", "mapinfo", "gene", "group", "relation", "table"])
    writer.writeheader()
    writer.writerows(essex_rows)
print(f"Wrote {ESSEX_HIGH_CORR_CSV} ({len(essex_rows):,} rows)")


# ===========================================================================
# STEP 3 — Extract Sommerer et al. 2022 buccal-brain correlations
# (Supplementary Table S6, the "top q<0.05" list), filter to |rho| > 0.6,
# pull CHR/MAPINFO from Sommerer's own reanalysis table (S2), and pull
# gene/genomic-context/CpG-island annotation from the SAME Essex Table 2
# manifest lookup built in Step 1 (Sommerer's tables don't carry gene
# annotation directly; EPIC probes largely overlap the 450K manifest).
# ===========================================================================
print("Loading Sommerer et al. 2022 buccal-brain tables from sommerer_supp1.xlsx ...")
wb_sommerer = openpyxl.load_workbook(SOMMERER_XLSX, read_only=True)

sommerer_chrpos_lookup = {}  # probe -> (chr, pos), from S2
ws_s2 = wb_sommerer["S2"]
for row in ws_s2.iter_rows(min_row=5, values_only=True):
    probe = row[0]
    if probe:
        sommerer_chrpos_lookup[probe] = (row[1], row[2])  # CHR, MAPINFO

buccal_rows = []
ws_s6 = wb_sommerer["S6"]
for row in ws_s6.iter_rows(min_row=3, values_only=True):
    probe = row[0]
    rho = row[1]
    if not probe or rho is None or abs(rho) <= CORR_THRESHOLD:
        continue
    q_val = row[3]
    buccal_mqtl, blood_mqtl, brain_mqtl, corsiv, braun = row[4], row[5], row[6], row[7], row[8]
    # Prefer Sommerer's own reanalysis table (S2) for CHR/MAPINFO; not every
    # S6 probe appears in S2 (S2 only covers CpGs previously published by
    # Braun et al.), so fall back to the shared Essex Table 2 manifest
    # lookup, which covers the full 450K array and catches most of the rest.
    chr_val, pos_val = sommerer_chrpos_lookup.get(probe, (None, None))
    if chr_val is None:
        chr_val, pos_val = chrpos_lookup.get(probe, (None, None))
    gene, group, relation = annotation_lookup.get(probe, ("", "", ""))
    buccal_rows.append({
        "probe": probe, "rho": rho, "q": q_val,
        "chr": chr_val, "pos": pos_val,
        "gene": gene, "group": group, "relation": relation,
        "buccal_mqtl": buccal_mqtl, "blood_mqtl": blood_mqtl,
        "brain_mqtl": brain_mqtl, "corsiv": corsiv,
        "in_braun_IMAGE_CpG": braun,
    })
print(f"  S6: {len(buccal_rows):,} qualifying CpGs with |rho| > {CORR_THRESHOLD}")

with open(BUCCAL_HIGH_CORR_CSV, "w", newline="", encoding="utf-8") as f:
    fieldnames = ["probe", "rho", "q", "chr", "pos", "gene", "group", "relation",
                  "buccal_mqtl", "blood_mqtl", "brain_mqtl", "corsiv", "in_braun_IMAGE_CpG"]
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(buccal_rows)
print(f"Wrote {BUCCAL_HIGH_CORR_CSV} ({len(buccal_rows):,} rows)")


# ===========================================================================
# STEP 4 — Merge both intermediate extracts into the master CSV, applying
# genomic-context classification, tissue-pair labels, correlation rounding,
# and per-source citation metadata.
# ===========================================================================
print("Merging into master CSV ...")
ESSEX_REGION_LABELS = {
    "PFC": "Blood vs PFC (brain)", "EC": "Blood vs EC (brain)",
    "STG": "Blood vs STG (brain)", "CER": "Blood vs CER (brain)",
}

master_rows = []
for r in essex_rows:
    master_rows.append({
        "cpg": r["probe"],
        "chr": r["chr"] if r["chr"] not in (None, "") else "",
        "pos": r["mapinfo"] if r["mapinfo"] not in (None, "") else "",
        "gene": r["gene"],
        "context": classify_context(r["group"]),
        "island_relation": r["relation"],
        "tissue_pair": ESSEX_REGION_LABELS[r["region"]],
        "corr_type": "Pearson r",
        "corr_value": round(float(r["r"]), 3),
        "source": HANNON_SOURCE,
        "source_url": HANNON_URL,
        "database_tool": HANNON_TOOL,
    })

for r in buccal_rows:
    master_rows.append({
        "cpg": r["probe"],
        "chr": r["chr"] if r["chr"] not in (None, "") else "",
        "pos": r["pos"] if r["pos"] not in (None, "") else "",
        "gene": r["gene"],
        "context": classify_context(r["group"]),
        "island_relation": r["relation"],
        "tissue_pair": "Buccal vs Prefrontal Cortex (brain)",
        "corr_type": "Spearman rho",
        "corr_value": round(float(r["rho"]), 3),
        "source": SOMMERER_SOURCE,
        "source_url": SOMMERER_URL,
        "database_tool": SOMMERER_TOOL,
    })

# NOTE: chr is stored as the RAW chromosome number here (e.g. "7", not "chr7").
# The "chr" prefix is added later by build_cpg_database.py when assembling the
# .xlsx workbook, matching the shipped database's CSV convention.
with open(MASTER_CSV, "w", newline="", encoding="utf-8") as f:
    fieldnames = ["cpg", "chr", "pos", "gene", "context", "island_relation", "tissue_pair",
                  "corr_type", "corr_value", "source", "source_url", "database_tool"]
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    for row in master_rows:
        writer.writerow(row)
print(f"Wrote {MASTER_CSV} ({len(master_rows):,} rows)")


# ===========================================================================
# STEP 5 — Build the gene-level summary: one row per unique gene symbol
# (genes are extracted from the semicolon-joined multi-transcript 'gene'
# field and deduplicated), aggregating qualifying CpG count, max |correlation|,
# tissue pairs covered, and genomic contexts covered.
# ===========================================================================
print("Building gene-level summary ...")


def split_genes(raw_gene):
    seen = []
    for part in raw_gene.split(";"):
        p = part.strip()
        if p and p not in seen:
            seen.append(p)
    return seen


gene_stats = {}  # gene -> {chr, cpgs:set, max_corr, tissue_pairs:set, contexts:set}
for row in master_rows:
    genes = split_genes(row["gene"])
    if not genes:
        continue
    for gene in genes:
        stats = gene_stats.setdefault(gene, {
            "chr": row["chr"], "cpgs": set(), "max_corr": 0.0,
            "tissue_pairs": set(), "contexts": set(),
        })
        stats["cpgs"].add(row["cpg"])
        stats["max_corr"] = max(stats["max_corr"], abs(row["corr_value"]))
        stats["tissue_pairs"].add(row["tissue_pair"])
        stats["contexts"].add(row["context"])

with open(GENE_SUMMARY_CSV, "w", newline="", encoding="utf-8") as f:
    fieldnames = ["gene", "chr", "n_qualifying_cpgs", "max_abs_corr", "tissue_pairs", "contexts"]
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    # Sort by qualifying CpG count descending, then gene name, to match the
    # ordering convention used in the shipped database.
    for gene, stats in sorted(gene_stats.items(), key=lambda kv: (-len(kv[1]["cpgs"]), kv[0])):
        writer.writerow({
            "gene": gene,
            "chr": stats["chr"],
            "n_qualifying_cpgs": len(stats["cpgs"]),
            "max_abs_corr": round(stats["max_corr"], 3),
            "tissue_pairs": "; ".join(sorted(stats["tissue_pairs"])),
            "contexts": "; ".join(sorted(stats["contexts"])),
        })
print(f"Wrote {GENE_SUMMARY_CSV} ({len(gene_stats):,} genes)")

print("\nDone. Next step: run build_cpg_database.py to assemble the formatted "
      "5-sheet .xlsx workbook from these two CSVs.")
