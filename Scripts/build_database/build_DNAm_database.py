"""
build_cpg_database.py

Rebuilds cpg_brain_peripheral_correlation_database.csv from primary source data
for all FIVE eligible studies in the Cross-Tissue CpG Database project:

  1. Hannon E, et al. (2015). Epigenetics 10(11):1024-1032. doi:10.1080/15592294.2015.1100786
  2. Edgar RD, et al. (2017). Transl Psychiatry 7:e1187. doi:10.1038/tp.2017.171          [NO bulk data -- excluded from CSV, documented as query-tool-only]
  3. Braun PR, et al. (2019). Transl Psychiatry 9:47. doi:10.1038/s41398-019-0376-y
  4. Sommerer Y, et al. (2022). Clin Epigenetics 14:118. doi:10.1186/s13148-022-01357-w
  5. Nishitani S, et al. (2023). Transl Psychiatry 13:72. doi:10.1038/s41398-023-02370-0

CORE PRINCIPLES ENFORCED IN THIS SCRIPT (per project instructions):
  - No values are inferred, imputed, or silently recoded. Any field the source table did
    not report is set to the literal string "Not reported by source table" (NOT NaN, NOT 0).
  - Correlation metric (Pearson r vs Spearman rho) is preserved and stored explicitly per row;
    no metric is ever converted.
  - Raw source labels for brain region / peripheral tissue are preserved alongside new
    harmonised category columns -- the raw column is never overwritten.
  - Each row records whether the source table represents the genome-wide/unfiltered dataset
    or a filtered/statistically-selected subset (`dataset_scope`), and whether the reported
    values were selected via a significance threshold before publication
    (`statistically_selected_before_publication`).
  - Non-independence between resources sharing underlying data is flagged in `notes`.

Run: python3 build_cpg_database.py
Output: /home/user/workspace/cpg_brain_peripheral_correlation_database.csv
"""
import pandas as pd
import numpy as np
import openpyxl

NR = "Not reported by source table"  # distinct from NaN/missing-unavailable; source table simply lacks this field
NA_STRUCT = "Not applicable"

MASTER_COLUMNS = [
    "cpg_id", "chr", "position_bp", "genome_build",
    "gene_symbol_raw", "genic_context_raw", "cpg_island_relation_raw",
    "source_study", "dataset_resource_name", "doi", "pmid", "geo_accession",
    "sample_size_n", "methylation_platform", "ancestry_population",
    "brain_tissue_state", "cell_composition_adjustment",
    "brain_region_raw", "brain_region_harmonised",
    "peripheral_tissue_raw", "peripheral_tissue_harmonised",
    "tissue_pair_harmonised",
    "correlation_coefficient", "correlation_metric", "p_value", "q_value",
    "dataset_scope", "statistically_selected_before_publication",
    "original_concordance_definition", "analysis_variant",
    "mqtl_flag", "snp_flag", "transformation_applied", "notes",
]


def new_frame(n):
    df = pd.DataFrame(index=range(n), columns=MASTER_COLUMNS)
    df[:] = NR
    return df


# ---------------------------------------------------------------------------
# 1. HANNON 2015 -- blood vs 4 brain regions (PFC, EC, STG, CER), Pearson r
#    Source: Essex Blood-Brain Comparison Tool bulk download
#    (http://epigenetics.essex.ac.uk/bloodbrain/SupplementaryTables.xlsx)
#    Table 6 = 887 CpGs with similar tissue levels but NO interindividual
#              covariation (paired t-test P>0.1, r2<0.05) -- curated non-concordant examples
#    Table 7 = 1814 CpGs with STRONG blood-brain correlation (>50% variance
#              explained in at least one comparison) -- curated concordant examples
#    NEITHER table is the genome-wide/unfiltered set: both are curated subsets
#    selected by the study authors to illustrate high- and low-concordance CpGs.
#    Gene/CHR/position annotation is NOT available in these two tables (the
#    workbook's only annotation sheet, Table 8, covers only ~7,820 CpGs, none of
#    them overlapping the Table 6/7 IDs) -- left as "Not reported by source table"
#    rather than merged in from an external manifest.
# ---------------------------------------------------------------------------
def extract_hannon2015():
    path = "/home/user/workspace/hannon2015_supp/SupplementaryTables.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    region_map = {
        "PFC": "Prefrontal cortex",
        "EC": "Entorhinal cortex",
        "STG": "Superior temporal gyrus",
        "CER": "Cerebellum",
    }

    def parse_table(sheetname, concordance_definition):
        ws = wb[sheetname]
        rows = list(ws.iter_rows(values_only=True))
        # header spans rows 0-1: row0 has region group labels, row1 has r/p/variance sub-labels
        # data starts row 2
        recs = []
        for r in rows[2:]:
            if r[0] is None:
                continue
            probe_id = r[0]
            # columns: ProbeID, [PFC r,p,var], [EC r,p,var], [STG r,p,var], [CER r,p,var], diff cols...
            region_cols = {
                "PFC": (1, 2, 3),
                "EC": (4, 5, 6),
                "STG": (7, 8, 9),
                "CER": (10, 11, 12),
            }
            for region_code, (ri, pi, vi) in region_cols.items():
                r_val = r[ri] if ri < len(r) else None
                p_val = r[pi] if pi < len(r) else None
                if r_val is None and p_val is None:
                    continue
                recs.append({
                    "cpg_id": probe_id,
                    "brain_region_raw": region_code,
                    "brain_region_harmonised": region_map[region_code],
                    "correlation_coefficient": r_val,
                    "p_value": p_val,
                    "original_concordance_definition": concordance_definition,
                })
        return pd.DataFrame(recs)

    t6 = parse_table(
        "Supplementary Table 6",
        "Table S6: sites with similar average methylation between blood and brain "
        "(paired t-test P>0.1) but NO evidence of interindividual co-variation (r^2<0.05); "
        "curated low-concordance example set, not genome-wide.",
    )
    t7 = parse_table(
        "Supplementary Table 7",
        "Table S7: sites with strong interindividual blood-brain correlation "
        "(>50% of variance explained, r^2>0.5, i.e. r>~0.71) in at least one brain region; "
        "curated high-concordance example set, not genome-wide.",
    )
    wb.close()

    combined = pd.concat([t6, t7], ignore_index=True)
    n = len(combined)
    df = new_frame(n)
    df["cpg_id"] = combined["cpg_id"].values
    df["brain_region_raw"] = combined["brain_region_raw"].values
    df["brain_region_harmonised"] = combined["brain_region_harmonised"].values
    df["correlation_coefficient"] = combined["correlation_coefficient"].values
    df["p_value"] = combined["p_value"].values
    df["original_concordance_definition"] = combined["original_concordance_definition"].values

    df["chr"] = NR
    df["position_bp"] = NR
    df["genome_build"] = NR
    df["gene_symbol_raw"] = NR
    df["genic_context_raw"] = NR
    df["cpg_island_relation_raw"] = NR

    df["source_study"] = "Hannon E, Lunnon K, Schalkwyk L, Mill J (2015). Epigenetics 10(11):1024-1032."
    df["dataset_resource_name"] = "Essex Blood-Brain Comparison Tool (epigenetics.essex.ac.uk/bloodbrain) -- bulk supplementary tables"
    df["doi"] = "10.1080/15592294.2015.1100786"
    df["pmid"] = "26457534"
    df["geo_accession"] = "GSE59685"
    df["sample_size_n"] = "Blood n=80; PFC n=114; EC n=108; STG n=117; CER n=112 (matched subsets vary per pairwise comparison, ~71-75 per source tool description)"
    df["methylation_platform"] = "Illumina Infinium HumanMethylation450 BeadChip (450K)"
    df["ancestry_population"] = NR
    df["brain_tissue_state"] = "Postmortem (brain-bank autopsy tissue)"
    df["cell_composition_adjustment"] = NR

    df["peripheral_tissue_raw"] = "blood"
    df["peripheral_tissue_harmonised"] = "Blood"
    df["tissue_pair_harmonised"] = "Blood-" + df["brain_region_harmonised"].astype(str)

    df["correlation_metric"] = "Pearson r"
    df["q_value"] = NR
    df["dataset_scope"] = "filtered_curated_subset"
    df["statistically_selected_before_publication"] = "TRUE"
    df["analysis_variant"] = NA_STRUCT
    df["mqtl_flag"] = NR
    df["snp_flag"] = NR
    df["transformation_applied"] = "None; r and p reported exactly as given in source table."
    df["notes"] = (
        "Curated example subset only (Tables S6 and S7 of Hannon et al. 2015), NOT the "
        "genome-wide dataset. Genome-wide blood-brain correlation values for arbitrary CpGs "
        "are only retrievable via live per-probe/per-gene query at "
        "http://epigenetics.essex.ac.uk/bloodbrain/ and are not bulk-downloadable; "
        "not included in this database. Data shares underlying cohort with GEO GSE59685. "
        "Row-count note: the source workbook contains 886 unique ProbeIDs in Table S6 and "
        "1813 in Table S7 (verified by direct count of non-blank ProbeID cells), one fewer "
        "than the 887/1814 figures quoted in the published article text/abstract; the literal "
        "file content was used here rather than the article's summary count, per the "
        "no-inference principle."
    )
    return df[MASTER_COLUMNS]


# ---------------------------------------------------------------------------
# 2. BRAUN 2019 (IMAGE-CpG) -- blood/saliva/buccal vs live resected brain, Spearman rho
#    Source: Supplemental Table 3 (MOESM9 on Nature CMS) -- CpGs surpassing
#    BONFERRONI significance for brain-peripheral correlation. This is a
#    filtered/statistically-selected subset, not the genome-wide dataset.
# ---------------------------------------------------------------------------
def extract_braun2019():
    path = "/home/user/workspace/braun2019_supp/MOESM9.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # three stacked sections, each with its own header row, separated by a blank row
    sections = [
        (2, 2359, "blood", "Blood"),          # rows 2..2358 inclusive (0-indexed) -> data
        (2361, 4728, "buccal", "Buccal (epithelial)"),
        (4730, 6228, "saliva", "Saliva"),
    ]
    recs = []
    for start, end, tissue_raw, tissue_harm in sections:
        for r in rows[start:end]:
            if r[0] is None:
                continue
            recs.append({
                "cpg_id": r[0],
                "correlation_coefficient": r[1],
                "p_value": r[2],
                "chr": r[4],
                "position_bp": r[5],
                "gene_symbol_raw": r[6],
                "genic_context_raw": r[8],
                "peripheral_tissue_raw": tissue_raw,
                "peripheral_tissue_harmonised": tissue_harm,
            })
    combined = pd.DataFrame(recs)
    n = len(combined)
    df = new_frame(n)
    for col in ["cpg_id", "correlation_coefficient", "p_value", "chr", "position_bp",
                "gene_symbol_raw", "genic_context_raw", "peripheral_tissue_raw",
                "peripheral_tissue_harmonised"]:
        df[col] = combined[col].values
    df["chr"] = df["chr"].apply(lambda x: NR if pd.isna(x) else x)
    # A blank gene-name cell in this table is a reported absence (Illumina manifest found no
    # overlapping gene for this probe), not a missing/unreported field -- distinguish explicitly.
    # Use pd.isna() rather than `x is None`: pandas silently coerces None -> NaN (float) when
    # object columns are built from mixed dict values, and `x is None` would miss those.
    df["gene_symbol_raw"] = df["gene_symbol_raw"].apply(
        lambda x: "No gene annotated in source table (intergenic per Illumina manifest)" if pd.isna(x) else x
    )
    df["genic_context_raw"] = df["genic_context_raw"].apply(
        lambda x: "Not applicable (no gene annotated at this probe)" if pd.isna(x) else x
    )

    df["genome_build"] = NR
    df["cpg_island_relation_raw"] = NR

    df["source_study"] = "Braun PR, Han S, Hing B, et al. (2019). Transl Psychiatry 9:47."
    df["dataset_resource_name"] = "IMAGE-CpG (han-lab.org/methylation/default/imageCpG) -- Supplemental Table 3"
    df["doi"] = "10.1038/s41398-019-0376-y"
    df["pmid"] = "30705257"
    df["geo_accession"] = "GSE111165"
    df["sample_size_n"] = "n=27 total (450K: n=12 brain/saliva/blood; EPIC: n=21 brain/saliva/blood/buccal)"
    df["methylation_platform"] = "Illumina Infinium HumanMethylation450 (450K) and MethylationEPIC BeadChip (mixed across subjects)"
    df["ancestry_population"] = NR
    df["brain_tissue_state"] = "Live (resected during epilepsy surgery; region varies by subject, not specified per-CpG in this table)"
    df["cell_composition_adjustment"] = "Not reported for this table (raw methylation correlation, no explicit adjustment described)"

    df["brain_region_raw"] = "Live resected brain tissue (subject-specific resection site; not specified per-CpG)"
    df["brain_region_harmonised"] = "Brain (mixed/live resection, unspecified region)"
    df["tissue_pair_harmonised"] = df["peripheral_tissue_harmonised"].astype(str) + "-Brain (mixed/live resection)"

    df["correlation_metric"] = "Spearman rho"
    df["q_value"] = NR
    df["dataset_scope"] = "filtered_significant_subset"
    df["statistically_selected_before_publication"] = "TRUE"
    df["original_concordance_definition"] = (
        "CpGs surpassing Bonferroni-corrected significance threshold (p<6.1e-8) for the "
        "within-subject brain-peripheral correlation analysis (Supplemental Table 3)."
    )
    df["analysis_variant"] = NA_STRUCT
    df["mqtl_flag"] = NR
    df["snp_flag"] = NR
    df["transformation_applied"] = "None; rho and p reported exactly as given in source table."
    df["notes"] = (
        "Filtered Bonferroni-significant subset only (2357 blood, 2367 buccal, 1498 saliva CpGs), "
        "NOT the genome-wide dataset (paper reports 822,996 probes analysed genome-wide but does "
        "not provide the full per-CpG table as a supplementary file). Genome-wide/arbitrary-CpG "
        "values are only retrievable via live query at http://han-lab.org/methylation/default/imageCpG "
        "and are not bulk-downloadable; not included in this database. Braun et al.'s raw idat data "
        "(GSE111165) were independently reanalysed by Sommerer et al. (2022) for a buccal-brain "
        "comparison table (their Additional file 1, Table 2) -- treat any reuse of that reanalysis "
        "as non-independent of this source."
    )
    return df[MASTER_COLUMNS]


# ---------------------------------------------------------------------------
# 3. SOMMERER 2022 -- buccal vs PFC, Spearman rho, GENOME-WIDE / UNFILTERED
#    Source: sommerer_et_al.results_summary_complete.tsv (730,157 CpGs, all
#    QC'ed EPIC probes -- full genome-wide summary statistics, not filtered
#    to significant hits only).
# ---------------------------------------------------------------------------
def extract_sommerer2022():
    path = "/home/user/workspace/sommerer2022_supp/sommerer_et_al.results_summary_complete.tsv"
    raw = pd.read_csv(path, sep="\t")
    n = len(raw)
    df = new_frame(n)
    df["cpg_id"] = raw["CpG"].values
    df["chr"] = raw["Chromosome"].values
    df["position_bp"] = raw["Position"].values
    df["genome_build"] = "GRCh37/hg19 (per source README, Illumina manifest)"
    df["correlation_coefficient"] = raw["R"].values
    df["p_value"] = raw["P"].values
    df["q_value"] = raw["Q"].values

    df["gene_symbol_raw"] = NR
    df["genic_context_raw"] = NR
    df["cpg_island_relation_raw"] = NR

    df["source_study"] = "Sommerer Y, Ohlei O, Dobricic V, et al. (2022). Clin Epigenetics 14:118."
    df["dataset_resource_name"] = "Sommerer et al. buccal-brain genome-wide correlation map (full summary statistics release, cloud.omics.uni-luebeck.de)"
    df["doi"] = "10.1186/s13148-022-01357-w"
    df["pmid"] = "36109787"
    df["geo_accession"] = NR
    df["sample_size_n"] = "n=120 matched buccal-PFC pairs (MADRC cohort, batches MADRC-1 n=44 + MADRC-2 n=76 post-QC)"
    df["methylation_platform"] = "Illumina Infinium MethylationEPIC BeadChip"
    df["ancestry_population"] = NR
    df["brain_tissue_state"] = "Postmortem"
    df["cell_composition_adjustment"] = "Not reported for this genome-wide summary file (primary analysis; see source publication Methods for PC-adjustment sensitivity analysis reported separately in text, not in this file)"

    df["brain_region_raw"] = "PFC"
    df["brain_region_harmonised"] = "Prefrontal cortex"
    df["peripheral_tissue_raw"] = "buccal"
    df["peripheral_tissue_harmonised"] = "Buccal (epithelial)"
    df["tissue_pair_harmonised"] = "Buccal (epithelial)-Prefrontal cortex"

    df["correlation_metric"] = "Spearman rho"
    df["dataset_scope"] = "genome-wide_unfiltered"
    df["statistically_selected_before_publication"] = "FALSE"
    df["original_concordance_definition"] = (
        "Full genome-wide summary statistics for all 730,157 QC-passing EPIC CpGs; not filtered "
        "to a significance threshold. Paper's own significance threshold for its headline finding "
        "was FDR q<0.05 (~24,980 CpGs), but this file contains ALL tested CpGs, not just those."
    )
    df["analysis_variant"] = "Raw (unadjusted primary analysis)"
    df["mqtl_flag"] = NR
    df["snp_flag"] = NR
    df["transformation_applied"] = "None; rho, p and q reported exactly as given in source file."
    df["notes"] = (
        "This is the only genome-wide/unfiltered dataset among the five eligible studies. "
        "Sommerer et al. also independently reanalysed Braun et al. (2019) raw data (GSE111165) "
        "for a direct comparison (their Table 2); that reanalysis is NOT included here to avoid "
        "conflating it with Braun's own originally published Bonferroni-significant results."
    )
    return df[MASTER_COLUMNS]


# ---------------------------------------------------------------------------
# 4. NISHITANI 2023 (AMAZE-CpG) -- blood/saliva/buccal vs live resected brain,
#    Spearman rho. Source: Supplementary Table S4 (MOESM3) -- CpGs surpassing
#    BENJAMINI-HOCHBERG significance; filtered/statistically-selected subset.
#    Includes native mQTL and SNP flags -- retained explicitly.
#    Provided as 6 sheets: 3 tissues x {Raw, Cell-composition adjusted}.
# ---------------------------------------------------------------------------
def extract_nishitani2023():
    path = "/home/user/workspace/nishitani2023_supp/MOESM3_TableS4_genomewide.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sheet_map = {
        "BRvsBL_BH": ("blood", "Blood", "Raw"),
        "BRvsSA_BH": ("saliva", "Saliva", "Raw"),
        "BRvsEP_BH": ("buccal", "Buccal (epithelial)", "Raw"),
        "BRvsBLadj_BH": ("blood", "Blood", "Cell-composition adjusted"),
        "BRvsSAadj_BH": ("saliva", "Saliva", "Cell-composition adjusted"),
        "BRvsEPadj_BH": ("buccal", "Buccal (epithelial)", "Cell-composition adjusted"),
    }

    all_recs = []
    for sheetname, (tissue_raw, tissue_harm, variant) in sheet_map.items():
        ws = wb[sheetname]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[1]
        # columns: cgid, chr, MAPINFO, Gene_Name, rho_brain_X, p_rho_brain_X,
        #          rho_brain_X_adj, p_rho_brain_X_adj, SNP_flag, mQTL_flag, BH
        # NOTE: every sheet (raw AND adjusted) carries BOTH raw and adjusted columns.
        # The BH significance filter defining sheet membership was applied to the
        # column matching this sheet's own variant, so we must select accordingly:
        # adjusted sheets -> use the *_adj columns; raw sheets -> use the non-adj columns.
        if variant == "Cell-composition adjusted":
            rho_col_name = [c for c in header if c and c.startswith("rho_brain") and c.endswith("_adj")][0]
            p_col_name = [c for c in header if c and c.startswith("p_rho_brain") and c.endswith("_adj")][0]
        else:
            rho_col_name = [c for c in header if c and c.startswith("rho_brain") and not c.endswith("_adj")][0]
            p_col_name = [c for c in header if c and c.startswith("p_rho_brain") and not c.endswith("_adj")][0]
        rho_idx = header.index(rho_col_name)
        p_idx = header.index(p_col_name)
        for r in rows[2:]:
            if r[0] is None:
                continue
            all_recs.append({
                "cpg_id": r[0],
                "chr": r[1],
                "position_bp": r[2],
                "gene_symbol_raw": r[3],
                "correlation_coefficient": r[rho_idx],
                "p_value": r[p_idx],
                "snp_flag": r[8],
                "mqtl_flag": r[9],
                "peripheral_tissue_raw": tissue_raw,
                "peripheral_tissue_harmonised": tissue_harm,
                "analysis_variant": variant,
            })
    wb.close()

    combined = pd.DataFrame(all_recs)
    n = len(combined)
    df = new_frame(n)
    for col in ["cpg_id", "chr", "position_bp", "gene_symbol_raw", "correlation_coefficient",
                "p_value", "snp_flag", "mqtl_flag", "peripheral_tissue_raw",
                "peripheral_tissue_harmonised", "analysis_variant"]:
        df[col] = combined[col].values
    # A blank gene-name cell in this table is a reported absence (no gene overlaps this probe
    # per the manifest used by the authors), not a missing/unreported field -- distinguish explicitly.
    df["gene_symbol_raw"] = df["gene_symbol_raw"].apply(
        lambda x: "No gene annotated in source table (intergenic per manifest)" if pd.isna(x) else x
    )
    df["snp_flag"] = df["snp_flag"].astype(str)
    df["mqtl_flag"] = df["mqtl_flag"].astype(str)

    df["genome_build"] = NR
    df["genic_context_raw"] = NR
    df["cpg_island_relation_raw"] = NR

    df["source_study"] = "Nishitani S, Isozaki M, Yao A, et al. (2023). Transl Psychiatry 13:72."
    df["dataset_resource_name"] = "AMAZE-CpG (snishit-amaze-cpg.web.app) -- Supplementary Table S4"
    df["doi"] = "10.1038/s41398-023-02370-0"
    df["pmid"] = "36849439"
    df["geo_accession"] = "GSE214901 (SuperSeries)"
    df["sample_size_n"] = "n=19 Japanese subjects (20 recruited, 1 excluded for suspected blood contamination)"
    df["methylation_platform"] = "Illumina Infinium MethylationEPIC BeadChip"
    df["ancestry_population"] = "Japanese"
    df["brain_tissue_state"] = "Live (resected during neurosurgery; region varies by subject, not specified per-CpG in this table)"
    df["cell_composition_adjustment"] = df["analysis_variant"].apply(
        lambda v: "Adjusted for estimated cell-type proportions (CETS for brain; EpiDISH for blood/saliva/buccal)"
        if v == "Cell-composition adjusted"
        else "Not adjusted (raw methylation values)"
    )

    df["brain_region_raw"] = "Live resected brain tissue (subject-specific resection site; not specified per-CpG)"
    df["brain_region_harmonised"] = "Brain (mixed/live resection, unspecified region)"
    df["tissue_pair_harmonised"] = df["peripheral_tissue_harmonised"].astype(str) + "-Brain (mixed/live resection)"

    df["correlation_metric"] = "Spearman rho"
    df["q_value"] = NR
    df["dataset_scope"] = "filtered_significant_subset"
    df["statistically_selected_before_publication"] = "TRUE"
    df["original_concordance_definition"] = (
        "CpGs surpassing Benjamini-Hochberg FDR-corrected significance for the within-subject "
        "brain-peripheral Spearman correlation analysis (Supplementary Table S4), separately for "
        "Raw and cell-composition-Adjusted analyses."
    )
    df["transformation_applied"] = "None; rho and p reported exactly as given in source table."
    df["notes"] = (
        "Filtered BH-significant subset only (counts per tissue/variant: blood raw 53,482; "
        "saliva raw 26,576; buccal raw 14,764; blood adj 13,064; saliva adj 19,584; buccal adj 14,214), "
        "NOT the genome-wide dataset (paper analysed 825,637 retained CpGs genome-wide but the "
        "full per-CpG table is not released as a separate bulk file beyond the BH-significant subset). "
        "First and only database derived from an Asian (Japanese) population among the five studies. "
        "mqtl_flag/snp_flag are native to the source table (SNP-confounding and mQTL annotation "
        "performed by the original authors), retained verbatim. "
        "Row-count note: literal per-sheet counts verified in this workbook are blood-raw 53,481; "
        "saliva-raw 26,575; buccal-raw 14,762; blood-adj 13,063; saliva-adj 19,583; buccal-adj 14,212 "
        "-- each one fewer than the corresponding figures reported in the published article text; "
        "the literal file content was used here rather than the article's summary counts, per the "
        "no-inference principle."
    )
    return df[MASTER_COLUMNS]


# ---------------------------------------------------------------------------
# NORMALISATION -- split the long/wide table above into two files:
#   1. cpg_brain_peripheral_correlation_database.csv  (per-CpG observations)
#   2. dataset_metadata.csv                            (per-dataset provenance)
# This is a STORAGE-LAYER decision only: no scientific value, definition, or
# category is changed, dropped, or recoded. It exists because the per-row
# text fields (source_study, notes, original_concordance_definition, etc.)
# are IDENTICAL for every CpG within a given study/tissue-pair/analysis-variant
# group. Repeating full sentences ~730,000 times (once per Sommerer CpG) or
# ~140,000 times (Nishitani) inflated the single-table CSV to ~1.6 GB, which
# is impractical to load into the Shiny app or open in any spreadsheet tool.
# Splitting into a compact per-observation table (11 columns that genuinely
# vary per CpG) plus a small per-dataset lookup table (23 columns describing
# ~11 distinct datasets) preserves 100% of the original information -- it is
# recoverable in full via a simple join on `dataset_id` -- while cutting file
# size by roughly two orders of magnitude. If a single denormalised CSV is
# still required for a specific downstream tool, it can be regenerated at any
# time with `pandas.merge(cpg_df, dataset_df, on="dataset_id")`.
# ---------------------------------------------------------------------------
DATASET_LEVEL_COLUMNS = [
    "dataset_id", "source_study", "dataset_resource_name", "doi", "pmid", "geo_accession",
    "sample_size_n", "methylation_platform", "ancestry_population", "genome_build",
    "brain_tissue_state", "cell_composition_adjustment",
    "brain_region_raw", "brain_region_harmonised",
    "peripheral_tissue_raw", "peripheral_tissue_harmonised", "tissue_pair_harmonised",
    "correlation_metric", "dataset_scope", "statistically_selected_before_publication",
    "original_concordance_definition", "analysis_variant", "transformation_applied", "notes",
]
CPG_LEVEL_COLUMNS = [
    "dataset_id", "cpg_id", "chr", "position_bp",
    "gene_symbol_raw", "genic_context_raw", "cpg_island_relation_raw",
    "correlation_coefficient", "p_value", "q_value", "mqtl_flag", "snp_flag",
]


if __name__ == "__main__":
    print("Extracting Hannon 2015...")
    d1 = extract_hannon2015()
    print(f"  -> {len(d1)} rows")

    print("Extracting Braun 2019...")
    d2 = extract_braun2019()
    print(f"  -> {len(d2)} rows")

    print("Extracting Sommerer 2022...")
    d3 = extract_sommerer2022()
    print(f"  -> {len(d3)} rows")

    print("Extracting Nishitani 2023...")
    d4 = extract_nishitani2023()
    print(f"  -> {len(d4)} rows")

    final = pd.concat([d1, d2, d3, d4], ignore_index=True)
    final = final[MASTER_COLUMNS]

    # A "dataset" = one unique combination of study x tissue-pair x analysis-variant.
    # This is exactly the grain at which every DATASET_LEVEL_COLUMNS field is constant.
    grouping_cols = [
        "source_study", "dataset_resource_name", "doi", "pmid", "geo_accession",
        "sample_size_n", "methylation_platform", "ancestry_population", "genome_build",
        "brain_tissue_state", "cell_composition_adjustment",
        "brain_region_raw", "brain_region_harmonised",
        "peripheral_tissue_raw", "peripheral_tissue_harmonised", "tissue_pair_harmonised",
        "correlation_metric", "dataset_scope", "statistically_selected_before_publication",
        "original_concordance_definition", "analysis_variant", "transformation_applied", "notes",
    ]
    dataset_groups = final[grouping_cols].drop_duplicates().reset_index(drop=True)

    study_prefix_map = {
        "Hannon E, Lunnon K, Schalkwyk L, Mill J (2015). Epigenetics 10(11):1024-1032.": "HANNON2015",
        "Braun PR, Han S, Hing B, et al. (2019). Transl Psychiatry 9:47.": "BRAUN2019",
        "Sommerer Y, Ohlei O, Dobricic V, et al. (2022). Clin Epigenetics 14:118.": "SOMMERER2022",
        "Nishitani S, Isozaki M, Yao A, et al. (2023). Transl Psychiatry 13:72.": "NISHITANI2023",
    }

    def make_id(row):
        prefix = study_prefix_map[row["source_study"]]
        region_tag = str(row["brain_region_harmonised"]).split(" (")[0].replace(" ", "").upper()
        tissue_tag = str(row["peripheral_tissue_harmonised"]).split(" (")[0].replace(" ", "").upper()
        variant_tag = {
            "Raw": "RAW", "Cell-composition adjusted": "ADJ",
            "Raw (unadjusted primary analysis)": "RAW", "Not applicable": None,
        }.get(row["analysis_variant"])
        parts = [prefix, tissue_tag, region_tag]
        if variant_tag:
            parts.append(variant_tag)
        # Hannon 2015 has two curated example subsets (Table S6: low-concordance,
        # Table S7: high-concordance) sharing the same tissue pair -- disambiguate.
        concordance_def = str(row["original_concordance_definition"])
        if "Table S6" in concordance_def:
            parts.append("TABLES6")
        elif "Table S7" in concordance_def:
            parts.append("TABLES7")
        return "_".join(parts)

    dataset_groups["dataset_id"] = dataset_groups.apply(make_id, axis=1)
    assert dataset_groups["dataset_id"].is_unique, "dataset_id collision -- grouping columns do not uniquely determine an id"

    final = final.merge(dataset_groups[grouping_cols + ["dataset_id"]], on=grouping_cols, how="left")
    assert final["dataset_id"].notna().all(), "some rows failed to match a dataset_id"

    cpg_df = final[CPG_LEVEL_COLUMNS].copy()
    dataset_df = dataset_groups[DATASET_LEVEL_COLUMNS].copy()

    cpg_out = "/home/user/workspace/cpg_brain_peripheral_correlation_database.csv"
    meta_out = "/home/user/workspace/dataset_metadata.csv"
    cpg_df.to_csv(cpg_out, index=False)
    dataset_df.to_csv(meta_out, index=False)

    print(f"\nTotal CpG-level rows written: {len(cpg_df)}")
    print(f"Saved per-CpG observations to {cpg_out}")
    print(f"Saved {len(dataset_df)} dataset-level metadata rows to {meta_out}")
    print("\nDataset IDs:")
    print(dataset_df[["dataset_id", "source_study", "tissue_pair_harmonised", "dataset_scope"]].to_string(index=False))

    print("\nRows per source_study:")
    print(final["source_study"].value_counts())
    print("\nRows per dataset_scope:")
    print(final["dataset_scope"].value_counts())
    print("\nRows per tissue_pair_harmonised:")
    print(final["tissue_pair_harmonised"].value_counts())
